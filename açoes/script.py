import pandas as pd
import numpy as np
from openpyxl import load_workbook

caminho_csv = "statusinvest-busca-avancada.csv"
df = pd.read_csv(caminho_csv)

# Especifique o caminho para o novo arquivo Excel
caminho_xlsx = "statusinvest-busca-avancada.xlsx"

# Use o método to_excel para escrever o DataFrame para um arquivo Excel
df.to_excel(caminho_xlsx, index=False)


book = load_workbook(caminho_xlsx)
writer = pd.ExcelWriter(caminho_xlsx, engine="openpyxl")
writer.book = book

df_filtrado = df[(df["DY"] >= 6) & (df["ROE"] >= 15)]
if "Filtrado_DY_ROE" in writer.book.sheetnames:
    del writer.book["Filtrado_DY_ROE"]

df_filtrado.to_excel(writer, sheet_name="Filtrado_DY_ROE", index=False)
writer.save()
writer.close()
